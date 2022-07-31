# Python program to read an excel file - Fetch all rows

# import openpyxl module
import openpyxl

# Give the location of the file
path = "D:\\C_Workspaces_Repositories\\GitHub_Repositories\\Python_Learning\\PDF & EXCEL\\example.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# printing the value of first column
# Loop will print all values
# of first column
print("\nValue of first column")
for i in range(1, row + 1):
    for j in range(1, column + 1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        print(cell_obj.value, end = " ")
    print()

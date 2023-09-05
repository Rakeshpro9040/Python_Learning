from openpyxl import load_workbook
from datetime import datetime

# Load the Excel workbook
workbook = load_workbook("schedule_end.xlsx")
sheet = workbook.active

# Loop through the rows and convert custom dates
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    custom_date = row[0]  # Assuming custom dates are in the first column
    
    try:
        formatted_date = datetime.strptime(custom_date, "%d-%b-%y %I.%M %p")
    except (ValueError, TypeError):
        # Skip this row if there's an issue with the value
        continue
    
    # Write the formatted date back to the sheet
    sheet.cell(row=row_index, column=2, value=formatted_date)  # Assuming you want to write it in the second column

# Save the modified workbook
workbook.save("schedule_end.xlsx")